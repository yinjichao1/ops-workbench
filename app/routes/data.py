"""Data entry API — weekly entry + CSV import."""

import csv
import io
from datetime import date, datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, File, UploadFile, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
from ..models import get_db
from ..models.platform_metrics import PlatformDailyMetrics

router = APIRouter()


def parse_week(w: str) -> date:
    """Parse '2026-W32' -> Monday date, or 'YYYY-MM-DD' -> date."""
    if "-W" in w:
        y, wk = w.split("-W")
        return date.fromisocalendar(int(y), int(wk), 1)
    return date.fromisoformat(w)


def week_to_str(d: date) -> str:
    """Convert date to 'YYYY-Www' format."""
    iso = d.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def week_to_cn(d: date) -> str:
    """Convert date to Chinese format: '2026年8月第2周'."""
    # Find first Monday of the month
    first_day = d.replace(day=1)
    first_monday = first_day + timedelta(days=(7 - first_day.weekday()) % 7)
    if first_monday.month != d.month:
        first_monday = d.replace(day=1)
        while first_monday.weekday() != 0:
            first_monday += timedelta(days=1)
    if d < first_monday:
        # Belongs to previous month's week
        prev_month = first_day - timedelta(days=1)
        prev_first = prev_month.replace(day=1)
        prev_first_mon = prev_first + timedelta(days=(7 - prev_first.weekday()) % 7)
        if prev_first_mon.month != prev_first.month:
            prev_first_mon = prev_first
            while prev_first_mon.weekday() != 0:
                prev_first_mon += timedelta(days=1)
        week_num = ((d - prev_first_mon).days // 7) + 1
        return f"{prev_first_mon.year}年{prev_first_mon.month}月第{week_num}周"
    week_num = ((d - first_monday).days // 7) + 1
    return f"{d.year}年{d.month}月第{week_num}周"


# ---------- Pydantic schemas ----------
class DailyMetricIn(BaseModel):
    week: str  # "2026-W32" or "YYYY-MM-DD"
    platform: str
    account: str = ""
    followers: int = 0
    likes: int = 0
    comments: int = 0
    shares: int = 0
    new_followers: int = 0
    publish_count: int = 0
    plays: int = 0
    completion_rate: float = 0.0
    ad_spend: float = 0.0
    reads: int = 0
    in_views: int = 0
    conversion_count: int = 0
    note_reads: int = 0
    bookmarks: int = 0
    hearts: int = 0
    engagement_rate: float = 0.0
    is_viral: int = 0


class BatchMetricsIn(BaseModel):
    records: list[DailyMetricIn]


# ---------- 平台字段规范化（服务端兜底，防止播放/阅读双写翻倍） ----------
def normalize_platform_fields(platform: str, data: dict) -> dict:
    """各平台只保留一个曝光口径字段，服务端强制规范化：
    抖音/视频号 → plays；公众号 → reads；小红书 → note_reads。
    其余口径字段一律清零，看板求和 plays+reads+note_reads 才不会翻倍。
    与前端版本无关，即使旧页面/旧客户端写入也不会产生双倍数据。"""
    if platform == "小红书":
        if data.get("note_reads"):
            data["plays"] = 0
        elif data.get("plays"):
            data["note_reads"] = data["plays"]
            data["plays"] = 0
        data["reads"] = 0
    elif platform == "公众号":
        if data.get("reads"):
            data["plays"] = 0
        elif data.get("plays"):
            data["reads"] = data["plays"]
            data["plays"] = 0
        data["note_reads"] = 0
    elif platform in ("抖音", "视频号"):
        if data.get("plays"):
            data["reads"] = 0
            data["note_reads"] = 0
        elif data.get("reads"):
            data["plays"] = data["reads"]
            data["reads"] = 0
        elif data.get("note_reads"):
            data["plays"] = data["note_reads"]
            data["note_reads"] = 0
    return data


# ---------- Manual entry ----------
@router.post("/metrics")
def create_metric(body: DailyMetricIn, db: Session = Depends(get_db)):
    """单条录入周数据。week 格式如 '2026-W32'。"""
    d = parse_week(body.week)
    values = normalize_platform_fields(body.platform, body.model_dump(exclude={"week", "platform", "account"}))
    existing = (
        db.query(PlatformDailyMetrics)
        .filter(
            PlatformDailyMetrics.date == d,
            PlatformDailyMetrics.platform == body.platform,
            PlatformDailyMetrics.account == (body.account or ""),
        )
        .first()
    )
    if existing:
        for k, v in values.items():
            setattr(existing, k, v)
        db.commit()
        return {"ok": True, "updated": True, "id": existing.id, "week": week_to_str(d)}

    record = PlatformDailyMetrics(date=d, platform=body.platform, account=body.account or "")
    for k, v in values.items():
        setattr(record, k, v)
    db.add(record)
    db.commit()
    db.refresh(record)
    return {"ok": True, "updated": False, "id": record.id, "week": week_to_str(d)}


@router.post("/metrics/batch")
def create_metrics_batch(body: BatchMetricsIn, db: Session = Depends(get_db)):
    """批量录入周数据。"""
    count = 0
    for item in body.records:
        d = parse_week(item.week)
        values = normalize_platform_fields(item.platform, item.model_dump(exclude={"week", "platform"}))
        existing = (
            db.query(PlatformDailyMetrics)
            .filter(
                PlatformDailyMetrics.date == d,
                PlatformDailyMetrics.platform == item.platform,
            )
            .first()
        )
        if existing:
            for k, v in values.items():
                setattr(existing, k, v)
        else:
            record = PlatformDailyMetrics(date=d, platform=item.platform)
            for k, v in values.items():
                setattr(record, k, v)
            db.add(record)
        count += 1
    db.commit()
    return {"ok": True, "count": count}


# ---------- CSV import ----------
@router.post("/metrics/import-csv")
def import_metrics_csv(
    file: UploadFile = File(...),
    platform: str = Query(...),
    db: Session = Depends(get_db),
):
    """从 CSV 导入日报数据。"""
    content = file.file.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    count = 0

    field_map = {
        "周": "week", "粉丝数": "followers", "点赞": "likes",
        "评论": "comments", "分享": "shares", "新增粉丝": "new_followers",
        "发布数": "publish_count", "播放量": "plays", "完播率": "completion_rate",
        "投流消耗": "ad_spend", "阅读量": "reads", "在看": "in_views",
        "转化数": "conversion_count", "笔记阅读量": "note_reads",
        "收藏": "bookmarks", "互动率": "engagement_rate", "爆款": "is_viral",
    }

    for row in reader:
        data = {"platform": platform}
        for cn, en in field_map.items():
            val = row.get(cn, "").strip()
            if not val:
                continue
            if en == "week":
                try:
                    data[en] = val
                except ValueError:
                    continue
            elif en in ("completion_rate", "engagement_rate", "ad_spend"):
                data[en] = float(val) if val else 0.0
            elif en == "is_viral":
                data[en] = 1 if val.lower() in ("1", "true", "yes", "是") else 0
            else:
                data[en] = int(float(val)) if val else 0

        if "week" not in data:
            continue

        data = normalize_platform_fields(platform, data)
        d = parse_week(data["week"])
        existing = (
            db.query(PlatformDailyMetrics)
            .filter(
                PlatformDailyMetrics.date == d,
                PlatformDailyMetrics.platform == platform,
            )
            .first()
        )
        if existing:
            for k, v in data.items():
                if k not in ("week", "platform"):
                    setattr(existing, k, v)
        else:
            try:
                db.add(PlatformDailyMetrics(**data))
            except Exception:
                continue
        count += 1

    db.commit()
    return {"ok": True, "imported": count}


# ---------- 月度表格导入（抖音/视频号/小红书） ----------
def _parse_any_date(val) -> Optional[date]:
    """解析多种日期格式：2026/08/31、2026-08-02、2026年08月31日，或 datetime 对象。"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日", "%Y-%m", "%Y/%m"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_number(val) -> Optional[float]:
    """解析数值：支持 '54.49%'、'31.18s'、'1,234' 等格式。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace(",", "")
    if not s:
        return None
    num = ""
    for ch in s:
        if ch.isdigit() or ch in ".-+":
            num += ch
        elif num:
            break
    try:
        return float(num)
    except ValueError:
        return None


def _read_table_rows(filename: str, raw: bytes):
    """读取上传文件为二维表（list[list]）。支持 CSV（utf-8-sig/gbk）与 XLSX。"""
    lower = (filename or "").lower()
    if lower.endswith(".csv") or lower.endswith(".txt"):
        text = None
        for enc in ("utf-8-sig", "gbk", "utf-16"):
            try:
                text = raw.decode(enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if text is None:
            raise ValueError("CSV 编码无法识别，请另存为 UTF-8 后重试")
        rows = list(csv.reader(io.StringIO(text)))
        return [r for r in rows if any(str(c).strip() for c in r)]
    if lower.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        out = []
        for ws in wb.worksheets:
            sheet_rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
            out.append((ws.title, [r for r in sheet_rows if any(c is not None and str(c).strip() for c in r)]))
        wb.close()
        return out  # xlsx 返回 [(sheet_title, rows), ...]
    raise ValueError("仅支持 .xlsx 或 .csv 文件（.xls 请先另存为 .xlsx）")


def _header_index(headers, keywords):
    """返回每个关键词在表头行中的列下标 dict；找不到的键省略。"""
    idx = {}
    for kw in keywords:
        for i, h in enumerate(headers):
            if h is not None and kw in str(h).strip():
                idx[kw] = i
                break
    return idx


def _parse_shipinhao(rows):
    """视频号视频详情 CSV：时间/播放/推荐/喜欢/评论/分享/关注 → 整月汇总。"""
    header_i = None
    for i, row in enumerate(rows[:6]):
        joined = "".join(str(c) for c in row if c is not None)
        if "时间" in joined and "播放" in joined:
            header_i = i
            break
    if header_i is None:
        raise ValueError("未找到表头行（需要包含 时间/播放 列）")
    col = _header_index(rows[header_i], ["时间", "播放", "喜欢", "评论", "分享", "关注"])
    if "播放" not in col or "时间" not in col:
        raise ValueError("视频号表缺少 时间/播放 列")
    agg, dates = {}, []
    for row in rows[header_i + 1:]:
        d = _parse_any_date(row[col["时间"]]) if col["时间"] < len(row) else None
        if not d:
            continue
        dates.append(d)
        for key, name in (("plays", "播放"), ("likes", "喜欢"), ("comments", "评论"), ("shares", "分享"), ("new_followers", "关注")):
            if name in col and col[name] < len(row):
                v = _to_number(row[col[name]])
                if v:
                    agg[key] = agg.get(key, 0) + v
    if not dates:
        raise ValueError("没有解析到有效数据行")
    return agg, dates


def _parse_douyin(rows):
    """抖音 exporter xlsx：日期/投稿量/总播放量/总点赞量/总分享量/总评论量/5秒完播率 → 整月汇总。"""
    header_i = None
    for i, row in enumerate(rows[:6]):
        joined = "".join(str(c) for c in row if c is not None)
        if "日期" in joined and "播放" in joined:
            header_i = i
            break
    if header_i is None:
        raise ValueError("未找到表头行（需要包含 日期/总播放量 列）")
    col = _header_index(rows[header_i], ["日期", "投稿量", "总播放量", "总点赞量", "总分享量", "总评论量", "5秒完播率"])
    if "总播放量" not in col or "日期" not in col:
        raise ValueError("抖音表缺少 日期/总播放量 列")
    agg, dates, cr_list = {}, [], []
    for row in rows[header_i + 1:]:
        d = _parse_any_date(row[col["日期"]]) if col["日期"] < len(row) else None
        if not d:
            continue
        dates.append(d)
        for key, name in (("publish_count", "投稿量"), ("plays", "总播放量"), ("likes", "总点赞量"), ("shares", "总分享量"), ("comments", "总评论量")):
            if name in col and col[name] < len(row):
                v = _to_number(row[col[name]])
                if v:
                    agg[key] = agg.get(key, 0) + v
        if "5秒完播率" in col and col["5秒完播率"] < len(row):
            v = _to_number(row[col["5秒完播率"]])
            if v is not None:
                cr_list.append(v)
    if not dates:
        raise ValueError("没有解析到有效数据行")
    if cr_list:
        agg["completion_rate"] = round(sum(cr_list) / len(cr_list), 2)
    return agg, dates


def _parse_xhs(sheets):
    """小红书 exporter xlsx 多 sheet：
    账号总体观看数据(指标/数值) → 观看=note_reads、总完播率=completion_rate；
    缺失时回退 观看趋势 sheet 求和。月份从趋势表日期识别。"""
    agg, dates = {}, []
    title_rows = None

    def read_kv_sheet(rows):
        data = {}
        for row in rows:
            if len(row) >= 2 and row[0] is not None:
                data[str(row[0]).strip()] = row[1]
        return data

    for title, rows in sheets:
        if "账号总体观看数据" in title and title_rows is None:
            title_rows = read_kv_sheet(rows)
        elif ("观看趋势" in title or "曝光趋势" in title) and not dates:
            header_i = 0
            col = _header_index(rows[header_i], ["日期", "数值"]) if rows else {}
            di, vi = col.get("日期", 0), col.get("数值", 1)
            for row in rows[header_i + 1:]:
                d = _parse_any_date(row[di]) if di < len(row) else None
                if d:
                    dates.append(d)

    if title_rows:
        note_reads = _to_number(title_rows.get("观看"))
        if note_reads:
            agg["note_reads"] = int(note_reads)
        cr = _to_number(title_rows.get("总完播率(%)") or title_rows.get("总完播率"))
        if cr is not None:
            agg["completion_rate"] = round(cr, 2)
    if "note_reads" not in agg and dates is not None:
        # 回退：观看趋势 sheet 逐日求和
        for title, rows in sheets:
            if "观看趋势" in title:
                col = _header_index(rows[0], ["日期", "数值"]) if rows else {}
                vi = col.get("数值", 1)
                total = 0
                for row in rows[1:]:
                    if vi < len(row):
                        v = _to_number(row[vi])
                        if v:
                            total += v
                if total:
                    agg["note_reads"] = int(total)
                break
    if not dates and title_rows is None:
        raise ValueError("未识别小红书数据表（需要 账号总体观看数据 或 观看趋势 sheet）")
    if not agg:
        raise ValueError("未解析到 观看/完播率 数值")
    return agg, dates


@router.post("/metrics/import-month")
def import_metrics_month(
    file: UploadFile = File(...),
    platform: str = Query(..., description="抖音/视频号/小红书"),
    account: str = Query("", description="归属账号，文件内不含账号时必填"),
    db: Session = Depends(get_db),
):
    """月度表格导入：解析平台导出的日粒度数据表，自动按月聚合成一条月度记录（date=当月1号）。
    仅覆盖表格中出现的字段，不影响已录入的其他字段（如粉丝数）。"""
    raw = file.file.read()
    filename = file.filename or ""
    parsed = _read_table_rows(filename, raw)

    if platform == "视频号":
        rows = parsed[0][1] if (isinstance(parsed, list) and parsed and isinstance(parsed[0], tuple)) else parsed
        agg, dates = _parse_shipinhao(rows)
    elif platform == "抖音":
        rows = parsed[0][1] if (isinstance(parsed, list) and parsed and isinstance(parsed[0], tuple)) else parsed
        agg, dates = _parse_douyin(rows)
    elif platform == "小红书":
        if not (isinstance(parsed, list) and parsed and isinstance(parsed[0], tuple)):
            raise ValueError("小红书数据请上传 .xlsx 文件")
        agg, dates = _parse_xhs(parsed)
    else:
        raise ValueError(f"平台 {platform} 暂不支持表格导入（支持抖音/视频号/小红书）")

    # 月份识别：取数据中出现最多的年月
    month_counter = {}
    for d in dates:
        key = f"{d.year}-{d.month:02d}"
        month_counter[key] = month_counter.get(key, 0) + 1
    month = max(month_counter, key=month_counter.get) if month_counter else ""
    if not month:
        raise ValueError("无法从表格识别月份，请检查日期列")
    y, m = int(month[:4]), int(month[5:7])
    record_date = date(y, m, 1)

    agg = {k: v for k, v in agg.items() if v is not None}
    agg = normalize_platform_fields(platform, {**{k: 0 for k in ("plays", "reads", "note_reads")}, **agg})
    fields = {k: int(round(v)) for k, v in agg.items() if k != "completion_rate" and isinstance(v, (int, float))}
    if isinstance(agg.get("completion_rate"), (int, float)):
        fields["completion_rate"] = float(agg["completion_rate"])

    existing = (
        db.query(PlatformDailyMetrics)
        .filter(
            PlatformDailyMetrics.date == record_date,
            PlatformDailyMetrics.platform == platform,
            PlatformDailyMetrics.account == (account or ""),
        )
        .first()
    )
    if existing:
        for k, v in fields.items():
            setattr(existing, k, v)
        target = existing
        updated = True
    else:
        target = PlatformDailyMetrics(date=record_date, platform=platform, account=account or "")
        for k, v in fields.items():
            setattr(target, k, v)
        db.add(target)
        updated = False
    db.commit()
    db.refresh(target)

    return {
        "ok": True,
        "platform": platform,
        "account": account or "",
        "month": month,
        "days": len(dates),
        "updated": updated,
        "fields": fields,
        "record_date": str(record_date),
    }


# ---------- Query ----------
@router.get("/metrics")
def list_metrics(
    platform: Optional[str] = None,
    account: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = Query(90, le=365),
    db: Session = Depends(get_db),
):
    """查询周数据。"""
    q = db.query(PlatformDailyMetrics)
    if platform:
        q = q.filter(PlatformDailyMetrics.platform == platform)
    if account:
        q = q.filter(PlatformDailyMetrics.account == account)
    if start_date:
        q = q.filter(PlatformDailyMetrics.date >= parse_week(start_date))
    if end_date:
        q = q.filter(PlatformDailyMetrics.date <= parse_week(end_date))
    rows = q.order_by(PlatformDailyMetrics.date.desc()).limit(limit).all()

    return {
        "data": [
            {
                "id": r.id,
                "week": week_to_str(r.date),
                "week_cn": week_to_cn(r.date),
                "date": str(r.date),
                "platform": r.platform,
                "account": r.account or "",
                "followers": r.followers,
                "likes": r.likes,
                "comments": r.comments,
                "shares": r.shares,
                "new_followers": r.new_followers,
                "publish_count": r.publish_count,
                "plays": r.plays,
                "completion_rate": r.completion_rate,
                "ad_spend": r.ad_spend,
                "reads": r.reads,
                "in_views": r.in_views,
                "conversion_count": r.conversion_count,
                "note_reads": r.note_reads,
                "bookmarks": r.bookmarks,
                "engagement_rate": r.engagement_rate,
                "is_viral": r.is_viral,
            }
            for r in rows
        ]
    }


# ---------- Account Config ----------
ACCOUNTS = {
    "抖音": ["思格电网", "安哥", "范校", "东北电气人都认"],
    "视频号": ["思格电网", "范校", "安哥"],
    "公众号": ["思格电网"],
    "小红书": ["小格", "学姐", "范校"],
}


@router.get("/accounts")
def list_accounts(platform: Optional[str] = None):
    """返回各平台账号列表。"""
    if platform:
        return {"accounts": ACCOUNTS.get(platform, ["主号"])}
    return {"accounts": ACCOUNTS}
